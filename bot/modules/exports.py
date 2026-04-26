import logging
from datetime import date, timedelta
from pathlib import Path

import config
from bot.db.database import db


def _period_dates(period: str) -> tuple[str | None, str | None]:
    today = date.today()
    if period == "today":
        return today.strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if period == "week":
        return (today - timedelta(days=7)).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if period == "month":
        return today.replace(day=1).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    if period == "year":
        return today.replace(month=1, day=1).strftime("%Y-%m-%d"), today.strftime("%Y-%m-%d")
    return None, None


async def export_to_excel(user_id: int, target: str, period: str = "month",
                          section_name: str = None) -> str | None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logging.error("openpyxl not installed")
        return None

    start_date, end_date = _period_dates(period)
    wb = openpyxl.Workbook()
    ws = wb.active

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F4F8F")
    header_align = Alignment(horizontal="center")

    def _header(ws, cols):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

    if target == "tasks":
        ws.title = "Задачи"
        _header(ws, ["#", "Задача", "Приоритет", "Дата", "Время", "Статус", "Создана"])
        tasks = await db.task_list(user_id, status="pending")
        done_tasks = await db.task_list(user_id, status="done")
        all_tasks = tasks + done_tasks
        for t in all_tasks:
            ws.append([t["id"], t["title"], t.get("priority", "normal"),
                       t.get("due_date", ""), t.get("due_time", ""),
                       t.get("status", ""), t.get("created_at", "")[:10]])

    elif target == "expenses":
        ws.title = "Расходы"
        _header(ws, ["#", "Дата", "Сумма", "Валюта", "Описание", "Проект"])
        expenses = await db.expense_stats(user_id, start_date=start_date, end_date=end_date)
        for e in expenses:
            ws.append([e["id"], e.get("date", ""), e["amount"], e.get("currency", "USD"),
                       e.get("description", ""), e.get("project_name", "")])
        # Totals
        ws.append([])
        ws.append(["", "ИТОГО:"])
        totals = await db.expenses_total(user_id, start_date=start_date, end_date=end_date)
        for t in totals:
            ws.append(["", "", t["total"], t["currency"], f"({t['count']} записей)"])

    elif target == "section" and section_name:
        section = await db.section_find(user_id, section_name)
        if not section:
            return None
        ws.title = section["title"][:31]
        fields = section["fields"]
        _header(ws, ["#", "Дата"] + fields)
        records = await db.section_records_all(user_id, section_name)
        for r in records:
            row = [r["id"], r.get("created_at", "")[:10]]
            for f in fields:
                row.append(r["data"].get(f, ""))
            ws.append(row)

    elif target == "all":
        # Multiple sheets
        wb.remove(ws)

        # Tasks sheet
        ws_tasks = wb.create_sheet("Задачи")
        _header(ws_tasks, ["#", "Задача", "Приоритет", "Дата", "Статус"])
        tasks = await db.task_list(user_id)
        for t in tasks:
            ws_tasks.append([t["id"], t["title"], t.get("priority", ""),
                             t.get("due_date", ""), t.get("status", "")])

        # Expenses sheet
        ws_exp = wb.create_sheet("Расходы")
        _header(ws_exp, ["#", "Дата", "Сумма", "Валюта", "Описание", "Проект"])
        expenses = await db.expense_stats(user_id)
        for e in expenses:
            ws_exp.append([e["id"], e.get("date", ""), e["amount"],
                          e.get("currency", ""), e.get("description", ""), e.get("project_name", "")])

        # Sections
        sections = await db.section_list(user_id)
        for section in sections[:10]:  # max 10 sheets
            ws_s = wb.create_sheet(section["title"][:31])
            fields = section["fields"]
            _header(ws_s, ["#", "Дата"] + fields)
            records = await db.section_records_all(user_id, section["name"])
            for r in records:
                row = [r["id"], r.get("created_at", "")[:10]]
                for f in fields:
                    row.append(r["data"].get(f, ""))
                ws_s.append(row)
    else:
        return None

    # Auto-width
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    ts = date.today().strftime("%Y%m%d")
    filename = f"tuntun_{target}_{ts}.xlsx"
    file_path = config.EXPORTS_DIR / filename
    wb.save(str(file_path))

    await db.export_log(user_id, f"excel_{target}", str(file_path),
                        {"period": period, "section": section_name})
    return str(file_path)


async def export_to_txt(user_id: int, target: str, period: str = "month") -> str | None:
    start_date, end_date = _period_dates(period)
    lines = [f"TUNTUN Export — {target} — {date.today().strftime('%d.%m.%Y')}", "=" * 50, ""]

    if target in ("tasks", "all"):
        tasks = await db.task_list(user_id)
        lines.append("=== ЗАДАЧИ ===")
        for t in tasks:
            priority = {"high": "🔴", "normal": "🟡", "low": "🟢"}.get(t.get("priority", "normal"), "")
            due = f" [{t['due_date']}]" if t.get("due_date") else ""
            lines.append(f"{priority} #{t['id']}: {t['title']}{due} ({t.get('status', '')})")
        lines.append("")

    if target in ("expenses", "all"):
        expenses = await db.expense_stats(user_id, start_date=start_date, end_date=end_date)
        lines.append("=== РАСХОДЫ ===")
        for e in expenses:
            lines.append(f"  {e.get('date', '')} | {e['amount']} {e.get('currency', 'USD')} | {e.get('description', '')} | {e.get('project_name', '')}")
        totals = await db.expenses_total(user_id, start_date=start_date, end_date=end_date)
        for t in totals:
            lines.append(f"  ИТОГО {t['currency']}: {t['total']:.2f} ({t['count']} записей)")
        lines.append("")

    if target == "all":
        sections = await db.section_list(user_id)
        for section in sections:
            lines.append(f"=== {section['title'].upper()} ===")
            records = await db.section_records_all(user_id, section["name"])
            for r in records:
                data_str = " | ".join(f"{k}: {v}" for k, v in r["data"].items() if v is not None)
                lines.append(f"  [{r.get('created_at', '')[:10]}] {data_str}")
            lines.append("")

    ts = date.today().strftime("%Y%m%d")
    filename = f"tuntun_{target}_{ts}.txt"
    file_path = config.EXPORTS_DIR / filename
    file_path.write_text("\n".join(lines), encoding="utf-8")

    await db.export_log(user_id, f"txt_{target}", str(file_path), {"period": period})
    return str(file_path)


async def handle_excel(user_id: int, params: dict, ai_response: str, **kwargs) -> str:
    target = params.get("target", "all")
    section_name = params.get("section_name")
    period = params.get("period", "month")

    file_path = await export_to_excel(user_id, target, period, section_name)
    if not file_path:
        return "❌ Нет данных для экспорта или раздел не найден"

    _period_labels = {"today": "сегодня", "week": "неделю", "month": "месяц", "year": "год", "all": "всё время"}
    _target_labels = {"tasks": "Задачи", "expenses": "Расходы", "section": section_name or "Раздел", "all": "Все данные"}
    period_str = _period_labels.get(period, period)
    target_str = _target_labels.get(target, target)
    summary = f"📊 Экспорт Excel: {target_str} за {period_str}\n"
    return f"{summary}__FILE__:{file_path}"


async def handle_txt(user_id: int, params: dict, ai_response: str, **kwargs) -> str:
    target = params.get("target", "all")
    period = params.get("period", "month")

    file_path = await export_to_txt(user_id, target, period)
    if not file_path:
        return "❌ Нет данных для экспорта"

    _period_labels = {"today": "сегодня", "week": "неделю", "month": "месяц", "year": "год", "all": "всё время"}
    _target_labels = {"tasks": "Задачи", "expenses": "Расходы", "all": "Все данные"}
    period_str = _period_labels.get(period, period)
    target_str = _target_labels.get(target, target)
    summary = f"📄 Экспорт TXT: {target_str} за {period_str}\n"
    return f"{summary}__FILE__:{file_path}"
